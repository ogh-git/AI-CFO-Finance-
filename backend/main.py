import io
import json
import logging
import os
import sqlite3
import ssl
import decimal
import zipfile
import hashlib
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Any, List, Optional

import asyncpg
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel

try:
    import bcrypt as _bcrypt
    from jose import JWTError, jwt as jose_jwt

    class _PwdCtx:
        def hash(self, secret: str) -> str:
            return _bcrypt.hashpw(secret.encode(), _bcrypt.gensalt()).decode()
        def verify(self, secret: str, hashed: str) -> bool:
            try:
                return _bcrypt.checkpw(secret.encode(), hashed.encode())
            except Exception:
                return False

    AUTH_OK = True
except ImportError:
    AUTH_OK = False

# ── optional heavy deps ────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers as xl_num
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    import anthropic as ant
    ANT_OK = True
except ImportError:
    ANT_OK = False

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
EXTERNAL_AUDITOR_SESSION_DAYS = int(os.getenv("EXTERNAL_AUDITOR_SESSION_DAYS", "90"))

# ── Audit module ───────────────────────────────────
from audit.db        import init_audit_db, get_audit_conn, append_audit_log, verify_audit_log
from audit.controls  import CONTROLS_MAP
from audit.sampling  import run_sampling
from audit.snapshot  import lock_tb_snapshot, get_post_lock_adjustments
from audit.models    import (
    CreateEngagementReq, CreatePBCItemReq, UpdatePBCStatusReq, BulkImportPBCReq,
    CreateFindingReq, UpdateFindingReq, SamplingRunReq, TBLockReq,
    RiskRegisterUpdateReq, SoDStatusReq, AuditPackageReq,
)

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

load_dotenv()

# ── Auth / Users ──────────────────────────────────────
USERS_DB   = os.getenv("USERS_DB", os.path.join(os.path.dirname(__file__), "data", "users.db"))
JWT_SECRET = os.getenv("JWT_SECRET", "cfo-dashboard-secret-change-in-prod")
JWT_ALG    = "HS256"
JWT_HOURS  = 24
pwd_ctx    = _PwdCtx() if AUTH_OK else None

def _uconn():
    os.makedirs(os.path.dirname(USERS_DB), exist_ok=True)
    return sqlite3.connect(USERS_DB, check_same_thread=False)

def _init_users():
    c = _uconn()
    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        email TEXT NOT NULL DEFAULT '',
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'viewer',
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    c.commit()
    if c.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0 and AUTH_OK:
        c.execute("INSERT INTO users (username,email,password_hash,role) VALUES (?,?,?,?)",
                  ("admin","admin@example.com", pwd_ctx.hash("admin123"), "admin"))
        c.commit()
        log.info("Default admin created — username: admin  password: admin123  (change it!)")
    c.close()

def _make_token(uid: int, username: str, role: str) -> str:
    now = datetime.utcnow()
    payload = {"sub": str(uid), "username": username, "role": role,
               "iat": int(now.timestamp()),
               "exp": now + timedelta(hours=JWT_HOURS)}
    return jose_jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)

def _decode_token(token: str) -> dict:
    return jose_jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])

def _caller(request: Request) -> dict:
    token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
    try:
        return _decode_token(token)
    except Exception:
        raise HTTPException(401, "Invalid token")

# ── company_ids helper ────────────────────────────────
def parse_ids(s: Optional[str]) -> Optional[list]:
    if not s: return None
    try:
        ids = [int(x.strip()) for x in s.split(',') if x.strip()]
        return ids if ids else None
    except (ValueError, AttributeError):
        return None

DB_HOST = os.getenv("DB_HOST", "ogh-live-pg.techleara.net")
DB_PORT = int(os.getenv("DB_PORT", "22345"))
DB_USER = os.getenv("DB_USER", "postgres")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")
VALID_DBS = ["ogh-live", "77asia", "seeenviro"]
DB_LABELS = {"ogh-live": "OGH Live", "77asia": "77 Asia", "seeenviro": "SEE Enviro"}

app = FastAPI(title="AI CFO Finance Dashboard API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

@app.on_event("startup")
def startup():
    _init_users()
    init_audit_db()

AUDIT_ROLES = {"internal_auditor", "external_auditor", "audit_admin"}
VALID_ROLES = {"admin", "viewer"} | AUDIT_ROLES


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    public = ("/api/auth/login", "/api/health")
    if any(request.url.path.startswith(p) for p in public):
        return await call_next(request)
    token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
    if not token:
        return JSONResponse({"detail": "Authentication required"}, status_code=401)
    try:
        claims = _decode_token(token)
    except Exception:
        return JSONResponse({"detail": "Invalid or expired token"}, status_code=401)

    role = claims.get("role", "viewer")
    path = request.url.path

    # external_auditor: only /api/audit/* allowed
    if role == "external_auditor" and not path.startswith("/api/audit") and not path.startswith("/api/auth"):
        return JSONResponse({"detail": "Access restricted to audit module"}, status_code=403)

    # external_auditor session age check
    if role == "external_auditor":
        issued = claims.get("iat") or 0
        age_days = (datetime.utcnow().timestamp() - issued) / 86400
        if age_days > EXTERNAL_AUDITOR_SESSION_DAYS:
            return JSONResponse({"detail": "Auditor session expired"}, status_code=401)

    return await call_next(request)


def _ssl_ctx() -> ssl.SSLContext:
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx


async def _conn(db: str) -> asyncpg.Connection:
    if db not in VALID_DBS:
        raise HTTPException(400, f"Unknown database: {db}")
    try:
        return await asyncpg.connect(
            host=DB_HOST, port=DB_PORT, user=DB_USER,
            password=DB_PASSWORD, database=db, ssl=_ssl_ctx(),
        )
    except Exception as exc:
        log.error("DB connect error [%s@%s:%s/%s]: %s", DB_USER, DB_HOST, DB_PORT, db, exc)
        raise HTTPException(503, detail=f"Cannot connect to database '{db}': {exc}")


def _clean(v: Any) -> Any:
    if isinstance(v, decimal.Decimal):
        return float(v)
    return v


def _row(row) -> dict:
    return {k: _clean(v) for k, v in dict(row).items()}


async def fetch(db: str, sql: str, *args) -> list[dict]:
    conn = await _conn(db)
    try:
        rows = await conn.fetch(sql, *args)
        return [_row(r) for r in rows]
    except Exception as exc:
        log.error("fetch error [%s]: %s", db, exc, exc_info=True)
        raise HTTPException(500, detail=str(exc))
    finally:
        await conn.close()


async def fetch_one(db: str, sql: str, *args) -> dict:
    conn = await _conn(db)
    try:
        row = await conn.fetchrow(sql, *args)
        return _row(row) if row else {}
    except Exception as exc:
        log.error("fetch_one error [%s]: %s", db, exc, exc_info=True)
        raise HTTPException(500, detail=str(exc))
    finally:
        await conn.close()


# ─────────────────────────────────────────────
# /api/health
# ─────────────────────────────────────────────
@app.get("/api/health")
async def health():
    return {"status": "ok"}


# ─────────────────────────────────────────────
# Auth endpoints
# ─────────────────────────────────────────────
class LoginReq(BaseModel):
    username: str
    password: str

class CreateUserReq(BaseModel):
    username: str
    email: str = ""
    password: str
    role: str = "viewer"

class UpdateUserReq(BaseModel):
    email: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None

def _get_user(username: str):
    c = _uconn()
    row = c.execute("SELECT id,username,email,password_hash,role,is_active FROM users WHERE username=?",
                    (username,)).fetchone()
    c.close()
    if not row: return None
    return {"id":row[0],"username":row[1],"email":row[2],"password_hash":row[3],"role":row[4],"is_active":bool(row[5])}

@app.post("/api/auth/login")
async def auth_login(req: LoginReq):
    if not AUTH_OK: raise HTTPException(501, "Auth libraries not installed")
    user = _get_user(req.username)
    if not user or not user["is_active"] or not pwd_ctx.verify(req.password, user["password_hash"]):
        raise HTTPException(401, "Invalid credentials")
    token = _make_token(user["id"], user["username"], user["role"])
    return {"token": token, "user": {k: user[k] for k in ("id","username","email","role")}}

@app.get("/api/auth/me")
async def auth_me(request: Request):
    return _caller(request)

@app.get("/api/auth/users")
async def auth_list_users(request: Request):
    caller = _caller(request)
    if caller.get("role") != "admin": raise HTTPException(403, "Admins only")
    c = _uconn()
    rows = c.execute("SELECT id,username,email,role,is_active,created_at FROM users ORDER BY id").fetchall()
    c.close()
    return {"data": [{"id":r[0],"username":r[1],"email":r[2],"role":r[3],"is_active":bool(r[4]),"created_at":r[5]} for r in rows]}

@app.post("/api/auth/users")
async def auth_create_user(req: CreateUserReq, request: Request):
    caller = _caller(request)
    if caller.get("role") != "admin": raise HTTPException(403, "Admins only")
    if not AUTH_OK: raise HTTPException(501, "Auth libraries not installed")
    try:
        c = _uconn()
        c.execute("INSERT INTO users (username,email,password_hash,role) VALUES (?,?,?,?)",
                  (req.username, req.email, pwd_ctx.hash(req.password), req.role))
        c.commit()
        uid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
        c.close()
        return {"id": uid, "username": req.username, "email": req.email, "role": req.role, "is_active": True}
    except sqlite3.IntegrityError:
        raise HTTPException(400, f"Username '{req.username}' already exists")

@app.patch("/api/auth/users/{uid}")
async def auth_update_user(uid: int, req: UpdateUserReq, request: Request):
    caller = _caller(request)
    if caller.get("role") != "admin": raise HTTPException(403, "Admins only")
    c = _uconn()
    if req.email    is not None: c.execute("UPDATE users SET email=? WHERE id=?",      (req.email, uid))
    if req.role     is not None: c.execute("UPDATE users SET role=? WHERE id=?",       (req.role, uid))
    if req.is_active is not None: c.execute("UPDATE users SET is_active=? WHERE id=?", (1 if req.is_active else 0, uid))
    if req.password is not None and AUTH_OK:
        c.execute("UPDATE users SET password_hash=? WHERE id=?", (pwd_ctx.hash(req.password), uid))
    c.commit(); c.close()
    return {"ok": True}

@app.delete("/api/auth/users/{uid}")
async def auth_delete_user(uid: int, request: Request):
    caller = _caller(request)
    if caller.get("role") != "admin": raise HTTPException(403, "Admins only")
    if str(uid) == str(caller.get("sub")): raise HTTPException(400, "Cannot delete yourself")
    c = _uconn(); c.execute("DELETE FROM users WHERE id=?", (uid,)); c.commit(); c.close()
    return {"ok": True}


# ─────────────────────────────────────────────
# /api/companies
# ─────────────────────────────────────────────
@app.get("/api/companies")
async def companies():
    return {"companies": [{"id": k, "name": v} for k, v in DB_LABELS.items()]}


# ─────────────────────────────────────────────
# /api/sub-companies
# ─────────────────────────────────────────────
_EXCLUDED_ENTITIES = {
    "TSH HOLIDAY HOMES RENTAL",
    "JORDAN PROMOTION HOUSE LLC",
    "C H A MARTIAL ARTS",
    "IMMERSEE HOLDING",
}

@app.get("/api/sub-companies")
async def sub_companies(db: str = Query("ogh-live")):
    rows = await fetch(db, "SELECT id, name FROM res_company ORDER BY name")
    filtered = [r for r in rows if r["name"].strip().upper() not in _EXCLUDED_ENTITIES]
    return {"data": filtered}


# ─────────────────────────────────────────────
# /api/kpis   – KPI summary cards
# ─────────────────────────────────────────────
@app.get("/api/kpis")
async def kpis(
    db: str = Query("ogh-live"),
    year: int = Query(None),
    month: int = Query(None),
    company_ids: str = Query(None),
):
    ids = parse_ids(company_ids)
    today = date.today()
    y = year or today.year
    m = month or today.month

    pnl_sql = """
        SELECT
            COALESCE(SUM(CASE
                WHEN aa.account_type IN ('income','income_other')
                     AND EXTRACT(MONTH FROM am.date)::int = $1
                     AND EXTRACT(YEAR  FROM am.date)::int = $2
                THEN aml.credit - aml.debit ELSE 0 END), 0) AS month_revenue,
            COALESCE(SUM(CASE
                WHEN aa.account_type IN ('expense','expense_depreciation','expense_direct_cost')
                     AND EXTRACT(MONTH FROM am.date)::int = $1
                     AND EXTRACT(YEAR  FROM am.date)::int = $2
                THEN aml.debit - aml.credit ELSE 0 END), 0) AS month_expense,
            COALESCE(SUM(CASE
                WHEN aa.account_type IN ('income','income_other')
                     AND EXTRACT(YEAR  FROM am.date)::int = $2
                     AND EXTRACT(MONTH FROM am.date)::int <= $1
                THEN aml.credit - aml.debit ELSE 0 END), 0) AS ytd_revenue,
            COALESCE(SUM(CASE
                WHEN aa.account_type IN ('expense','expense_depreciation','expense_direct_cost')
                     AND EXTRACT(YEAR  FROM am.date)::int = $2
                     AND EXTRACT(MONTH FROM am.date)::int <= $1
                THEN aml.debit - aml.credit ELSE 0 END), 0) AS ytd_expense
        FROM account_move_line aml
        JOIN account_move    am ON am.id = aml.move_id
        JOIN account_account aa ON aa.id = aml.account_id
        WHERE am.state = 'posted'
          AND aa.account_type IN (
              'income','income_other',
              'expense','expense_depreciation','expense_direct_cost')
          AND EXTRACT(YEAR FROM am.date)::int = $2
          AND ($3::int[] IS NULL OR am.company_id = ANY($3::int[]))
    """
    ar_sql = """
        SELECT
            COALESCE(SUM(am.amount_residual), 0) AS total_ar,
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due > 30
                         THEN am.amount_residual ELSE 0 END), 0) AS overdue_ar
        FROM account_move am
        WHERE am.state = 'posted'
          AND am.move_type IN ('out_invoice','out_refund')
          AND am.amount_residual > 0
          AND ($1::int[] IS NULL OR am.company_id = ANY($1::int[]))
    """
    ap_sql = """
        SELECT
            COALESCE(SUM(am.amount_residual), 0) AS total_ap,
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due > 30
                         THEN am.amount_residual ELSE 0 END), 0) AS overdue_ap
        FROM account_move am
        WHERE am.state = 'posted'
          AND am.move_type IN ('in_invoice','in_refund')
          AND am.amount_residual > 0
          AND ($1::int[] IS NULL OR am.company_id = ANY($1::int[]))
    """
    pnl = await fetch_one(db, pnl_sql, m, y, ids)
    ar  = await fetch_one(db, ar_sql, ids)
    ap  = await fetch_one(db, ap_sql, ids)

    mr, me = float(pnl.get("month_revenue") or 0), float(pnl.get("month_expense") or 0)
    yr, ye = float(pnl.get("ytd_revenue") or 0),   float(pnl.get("ytd_expense") or 0)

    return {
        "period": {"year": y, "month": m},
        "month_revenue": mr,
        "month_expense": me,
        "month_profit":  mr - me,
        "month_margin":  round((mr - me) / mr * 100, 1) if mr else 0,
        "ytd_revenue":   yr,
        "ytd_expense":   ye,
        "ytd_profit":    yr - ye,
        "ytd_margin":    round((yr - ye) / yr * 100, 1) if yr else 0,
        "total_ar":      float(ar.get("total_ar") or 0),
        "overdue_ar":    float(ar.get("overdue_ar") or 0),
        "total_ap":      float(ap.get("total_ap") or 0),
        "overdue_ap":    float(ap.get("overdue_ap") or 0),
    }


# ─────────────────────────────────────────────
# /api/monthly-pnl   – Query 11
# ─────────────────────────────────────────────
@app.get("/api/monthly-pnl")
async def monthly_pnl(db: str = Query("ogh-live"), company_ids: str = Query(None)):
    sql = """
        SELECT
            TO_CHAR(am.date, 'YYYY-MM')            AS year_month,
            EXTRACT(YEAR  FROM am.date)::int        AS year,
            EXTRACT(MONTH FROM am.date)::int        AS month,
            COALESCE(SUM(CASE WHEN aa.account_type IN ('income','income_other')
                THEN aml.credit - aml.debit ELSE 0 END), 0) AS total_revenue,
            COALESCE(SUM(CASE WHEN aa.account_type IN (
                'expense','expense_depreciation','expense_direct_cost')
                THEN aml.debit - aml.credit ELSE 0 END), 0) AS total_expense
        FROM account_move_line aml
        JOIN account_move    am ON am.id = aml.move_id
        JOIN account_account aa ON aa.id = aml.account_id
        WHERE am.state = 'posted'
          AND am.date IS NOT NULL
          AND aa.account_type IN (
              'income','income_other',
              'expense','expense_depreciation','expense_direct_cost')
          AND am.date >= (CURRENT_DATE - INTERVAL '12 months')::date
          AND ($1::int[] IS NULL OR am.company_id = ANY($1::int[]))
        GROUP BY TO_CHAR(am.date, 'YYYY-MM'),
                 EXTRACT(YEAR  FROM am.date)::int,
                 EXTRACT(MONTH FROM am.date)::int
        ORDER BY year_month
    """
    ids = parse_ids(company_ids)
    rows = await fetch(db, sql, ids)
    return {"data": [
        {**r,
         "total_revenue": float(r["total_revenue"] or 0),
         "total_expense": float(r["total_expense"] or 0),
         "net_profit": float(r["total_revenue"] or 0) - float(r["total_expense"] or 0)}
        for r in rows
    ]}


# ─────────────────────────────────────────────
# /api/ar-aging   – Query 9 (summary)
# ─────────────────────────────────────────────
@app.get("/api/ar-aging")
async def ar_aging(db: str = Query("ogh-live"), company_ids: str = Query(None)):
    sql = """
        SELECT
            COALESCE(SUM(CASE WHEN am.invoice_date_due >= CURRENT_DATE
                              OR am.invoice_date_due IS NULL
                         THEN am.amount_residual ELSE 0 END), 0) AS current_bucket,
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 1  AND 30
                         THEN am.amount_residual ELSE 0 END), 0) AS bucket_1_30,
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 31 AND 60
                         THEN am.amount_residual ELSE 0 END), 0) AS bucket_31_60,
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 61 AND 90
                         THEN am.amount_residual ELSE 0 END), 0) AS bucket_61_90,
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due > 90
                         THEN am.amount_residual ELSE 0 END), 0) AS bucket_over_90,
            COALESCE(SUM(am.amount_residual), 0)                  AS total
        FROM account_move am
        WHERE am.state = 'posted'
          AND am.move_type IN ('out_invoice','out_refund')
          AND am.amount_residual > 0
          AND ($1::int[] IS NULL OR am.company_id = ANY($1::int[]))
    """
    ids = parse_ids(company_ids)
    row = await fetch_one(db, sql, ids)
    return {k: float(v or 0) for k, v in row.items()}


# ─────────────────────────────────────────────
# /api/ap-aging   – Query 10 (summary)
# ─────────────────────────────────────────────
@app.get("/api/ap-aging")
async def ap_aging(db: str = Query("ogh-live"), company_ids: str = Query(None)):
    sql = """
        SELECT
            COALESCE(SUM(CASE WHEN am.invoice_date_due >= CURRENT_DATE
                              OR am.invoice_date_due IS NULL
                         THEN am.amount_residual ELSE 0 END), 0) AS current_bucket,
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 1  AND 30
                         THEN am.amount_residual ELSE 0 END), 0) AS bucket_1_30,
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 31 AND 60
                         THEN am.amount_residual ELSE 0 END), 0) AS bucket_31_60,
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 61 AND 90
                         THEN am.amount_residual ELSE 0 END), 0) AS bucket_61_90,
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due > 90
                         THEN am.amount_residual ELSE 0 END), 0) AS bucket_over_90,
            COALESCE(SUM(am.amount_residual), 0)                  AS total
        FROM account_move am
        WHERE am.state = 'posted'
          AND am.move_type IN ('in_invoice','in_refund')
          AND am.amount_residual > 0
          AND ($1::int[] IS NULL OR am.company_id = ANY($1::int[]))
    """
    ids = parse_ids(company_ids)
    row = await fetch_one(db, sql, ids)
    return {k: float(v or 0) for k, v in row.items()}


# ─────────────────────────────────────────────
# /api/ar-customers  – Query 9 (detail)
# ─────────────────────────────────────────────
@app.get("/api/ar-customers")
async def ar_customers(db: str = Query("ogh-live"), limit: int = Query(25), company_ids: str = Query(None)):
    sql = """
        SELECT
            rp.name AS customer,
            COALESCE(SUM(CASE WHEN am.invoice_date_due >= CURRENT_DATE
                              OR am.invoice_date_due IS NULL
                         THEN am.amount_residual ELSE 0 END), 0) AS "Current",
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 1  AND 30
                         THEN am.amount_residual ELSE 0 END), 0) AS "1-30 Days",
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 31 AND 60
                         THEN am.amount_residual ELSE 0 END), 0) AS "31-60 Days",
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 61 AND 90
                         THEN am.amount_residual ELSE 0 END), 0) AS "61-90 Days",
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due > 90
                         THEN am.amount_residual ELSE 0 END), 0) AS "Over 90 Days",
            COALESCE(SUM(am.amount_residual), 0)                  AS total_outstanding
        FROM account_move am
        JOIN res_partner rp ON rp.id = am.partner_id
        WHERE am.state = 'posted'
          AND am.move_type IN ('out_invoice','out_refund')
          AND am.amount_residual > 0
          AND ($2::int[] IS NULL OR am.company_id = ANY($2::int[]))
        GROUP BY rp.name
        ORDER BY total_outstanding DESC
        LIMIT $1
    """
    ids = parse_ids(company_ids)
    rows = await fetch(db, sql, limit, ids)
    return {"data": rows}


# ─────────────────────────────────────────────
# /api/ap-vendors  – Query 10 (detail)
# ─────────────────────────────────────────────
@app.get("/api/ap-vendors")
async def ap_vendors(db: str = Query("ogh-live"), limit: int = Query(25), company_ids: str = Query(None)):
    sql = """
        SELECT
            rp.name AS vendor,
            COALESCE(SUM(CASE WHEN am.invoice_date_due >= CURRENT_DATE
                              OR am.invoice_date_due IS NULL
                         THEN am.amount_residual ELSE 0 END), 0) AS "Current",
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 1  AND 30
                         THEN am.amount_residual ELSE 0 END), 0) AS "1-30 Days",
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 31 AND 60
                         THEN am.amount_residual ELSE 0 END), 0) AS "31-60 Days",
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 61 AND 90
                         THEN am.amount_residual ELSE 0 END), 0) AS "61-90 Days",
            COALESCE(SUM(CASE WHEN CURRENT_DATE - am.invoice_date_due > 90
                         THEN am.amount_residual ELSE 0 END), 0) AS "Over 90 Days",
            COALESCE(SUM(am.amount_residual), 0)                  AS total_outstanding
        FROM account_move am
        JOIN res_partner rp ON rp.id = am.partner_id
        WHERE am.state = 'posted'
          AND am.move_type IN ('in_invoice','in_refund')
          AND am.amount_residual > 0
          AND ($2::int[] IS NULL OR am.company_id = ANY($2::int[]))
        GROUP BY rp.name
        ORDER BY total_outstanding DESC
        LIMIT $1
    """
    ids = parse_ids(company_ids)
    rows = await fetch(db, sql, limit, ids)
    return {"data": rows}


# ─────────────────────────────────────────────
# /api/pnl-detail  – Query 3
# ─────────────────────────────────────────────
@app.get("/api/pnl-detail")
async def pnl_detail(
    db: str = Query("ogh-live"),
    year: int = Query(None),
    month: int = Query(None),
    company_ids: str = Query(None),
):
    ids = parse_ids(company_ids)
    today = date.today()
    y = year or today.year
    m = month or today.month
    sql = """
        SELECT
            CASE
                WHEN aa.account_type IN ('income','income_other') THEN 'Revenue'
                WHEN aa.account_type = 'expense_direct_cost'      THEN 'Cost of Sales'
                WHEN aa.account_type = 'expense_depreciation'     THEN 'Depreciation'
                WHEN aa.account_type = 'expense'                  THEN 'Operating Expense'
                ELSE aa.account_type
            END AS category,
            aa.code_store->>(aml.company_id::text) AS account_code,
            aa.name->>'en_US'                       AS account_name,
            COALESCE(SUM(
                CASE WHEN aa.account_type IN ('income','income_other')
                     THEN aml.credit - aml.debit
                     ELSE aml.debit - aml.credit END
            ), 0) AS amount
        FROM account_move_line aml
        JOIN account_move    am ON am.id = aml.move_id
        JOIN account_account aa ON aa.id = aml.account_id
        WHERE am.state = 'posted'
          AND aa.account_type IN (
              'income','income_other',
              'expense','expense_depreciation','expense_direct_cost')
          AND EXTRACT(YEAR  FROM am.date)::int = $1
          AND EXTRACT(MONTH FROM am.date)::int = $2
          AND ($3::int[] IS NULL OR am.company_id = ANY($3::int[]))
        GROUP BY aa.account_type,
                 aa.code_store->>(aml.company_id::text),
                 aa.name->>'en_US'
        ORDER BY category, account_code
    """
    rows = await fetch(db, sql, y, m, ids)
    return {"data": rows, "period": {"year": y, "month": m}}


# ─────────────────────────────────────────────
# /api/balance-sheet  – Query 4
# ─────────────────────────────────────────────
@app.get("/api/balance-sheet")
async def balance_sheet(db: str = Query("ogh-live"), company_ids: str = Query(None)):
    sql = """
        SELECT
            CASE
                WHEN aa.account_type LIKE 'asset%'         THEN 'Asset'
                WHEN aa.account_type LIKE 'liability%'     THEN 'Liability'
                WHEN aa.account_type = 'equity'            THEN 'Equity'
                WHEN aa.account_type = 'equity_unaffected' THEN 'Retained Earnings'
                ELSE aa.account_type
            END AS category,
            aa.code_store->>(aml.company_id::text) AS account_code,
            aa.name->>'en_US'                       AS account_name,
            COALESCE(SUM(aml.balance), 0)           AS net_balance
        FROM account_move_line aml
        JOIN account_move    am ON am.id = aml.move_id
        JOIN account_account aa ON aa.id = aml.account_id
        WHERE am.state = 'posted'
          AND aa.account_type IN (
              'asset_receivable','asset_cash','asset_current',
              'asset_non_current','asset_prepayments','asset_fixed',
              'liability_payable','liability_current','liability_non_current',
              'equity','equity_unaffected')
          AND ($1::int[] IS NULL OR am.company_id = ANY($1::int[]))
        GROUP BY aa.account_type,
                 aa.code_store->>(aml.company_id::text),
                 aa.name->>'en_US'
        ORDER BY category, account_code
    """
    ids = parse_ids(company_ids)
    rows = await fetch(db, sql, ids)
    return {"data": rows}


# ─────────────────────────────────────────────
# /api/trial-balance  – Query 2
# ─────────────────────────────────────────────
@app.get("/api/trial-balance")
async def trial_balance(db: str = Query("ogh-live"), company_ids: str = Query(None)):
    sql = """
        SELECT
            aa.code_store->>(aml.company_id::text) AS account_code,
            aa.name->>'en_US'                       AS account_name,
            aa.account_type,
            COALESCE(SUM(aml.debit),   0)           AS total_debit,
            COALESCE(SUM(aml.credit),  0)           AS total_credit,
            COALESCE(SUM(aml.balance), 0)           AS net_balance
        FROM account_move_line aml
        JOIN account_move    am ON am.id = aml.move_id
        JOIN account_account aa ON aa.id = aml.account_id
        WHERE am.state = 'posted'
          AND ($1::int[] IS NULL OR aml.company_id = ANY($1::int[]))
        GROUP BY aa.code_store->>(aml.company_id::text),
                 aa.name->>'en_US', aa.account_type
        ORDER BY account_code
    """
    ids = parse_ids(company_ids)
    rows = await fetch(db, sql, ids)
    return {"data": rows}


# ─────────────────────────────────────────────
# /api/cash  – Query 7
# ─────────────────────────────────────────────
@app.get("/api/cash")
async def cash(db: str = Query("ogh-live"), limit: int = Query(100),
               company_ids: str = Query(None),
               year: str = Query(None), month: str = Query(None)):
    sql = """
        SELECT
            am.name                 AS payment_reference,
            am.date                 AS payment_date,
            aj.code                 AS journal_code,
            aj.name->>'en_US'       AS journal_name,
            aj.type                 AS journal_type,
            rp.name                 AS partner,
            rc.name                 AS currency,
            COALESCE(am.amount_total, 0) AS amount,
            am.move_type,
            am.state
        FROM account_move am
        JOIN account_journal aj ON aj.id = am.journal_id
        JOIN res_company     co ON co.id = am.company_id
        JOIN res_currency    rc ON rc.id = am.currency_id
        LEFT JOIN res_partner rp ON rp.id = am.partner_id
        WHERE am.state = 'posted'
          AND aj.type IN ('bank','cash')
          AND ($2::int[] IS NULL OR am.company_id = ANY($2::int[]))
          AND ($3::int[] IS NULL OR EXTRACT(YEAR  FROM am.date)::int = ANY($3::int[]))
          AND ($4::int[] IS NULL OR EXTRACT(MONTH FROM am.date)::int = ANY($4::int[]))
        ORDER BY am.date DESC
        LIMIT $1
    """
    ids    = parse_ids(company_ids)
    years  = parse_ids(year)
    months = parse_ids(month)
    rows = await fetch(db, sql, limit, ids, years, months)
    return {"data": rows}


# ─────────────────────────────────────────────
# /api/invoices  – Query 5
# ─────────────────────────────────────────────
@app.get("/api/invoices")
async def invoices(db: str = Query("ogh-live"), limit: int = Query(100),
                   company_ids: str = Query(None),
                   year: str = Query(None), month: str = Query(None)):
    sql = """
        SELECT
            am.name                 AS invoice_number,
            am.invoice_date         AS invoice_date,
            am.invoice_date_due     AS due_date,
            am.move_type,
            am.state,
            am.payment_state,
            rp.name                 AS customer,
            rc.name                 AS currency,
            COALESCE(am.amount_total,    0) AS total_amount,
            COALESCE(am.amount_residual, 0) AS amount_due,
            CASE
                WHEN am.amount_residual = 0                               THEN 'Paid'
                WHEN am.invoice_date_due IS NULL                          THEN 'No Due Date'
                WHEN am.invoice_date_due >= CURRENT_DATE                  THEN 'Current'
                WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 1  AND 30 THEN '1-30 Days'
                WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 31 AND 60 THEN '31-60 Days'
                WHEN CURRENT_DATE - am.invoice_date_due BETWEEN 61 AND 90 THEN '61-90 Days'
                ELSE 'Over 90 Days'
            END AS aging_bucket,
            GREATEST(0, CURRENT_DATE - am.invoice_date_due) AS days_overdue
        FROM account_move am
        JOIN res_partner  rp ON rp.id = am.partner_id
        JOIN res_currency rc ON rc.id = am.currency_id
        WHERE am.state = 'posted'
          AND am.move_type IN ('out_invoice','out_refund')
          AND ($2::int[] IS NULL OR am.company_id = ANY($2::int[]))
          AND ($3::int[] IS NULL OR EXTRACT(YEAR  FROM am.invoice_date)::int = ANY($3::int[]))
          AND ($4::int[] IS NULL OR EXTRACT(MONTH FROM am.invoice_date)::int = ANY($4::int[]))
        ORDER BY am.invoice_date DESC
        LIMIT $1
    """
    ids    = parse_ids(company_ids)
    years  = parse_ids(year)
    months = parse_ids(month)
    rows = await fetch(db, sql, limit, ids, years, months)
    return {"data": rows}


# ─────────────────────────────────────────────
# /api/partners  – Partner Master
# ─────────────────────────────────────────────
@app.get("/api/partners")
async def partners(db: str = Query("ogh-live"), limit: int = Query(200)):
    sql = """
        SELECT
            rp.name                                     AS partner_name,
            CASE WHEN rp.customer_rank > 0 AND rp.supplier_rank > 0
                      THEN 'Customer & Vendor'
                 WHEN rp.customer_rank > 0 THEN 'Customer'
                 WHEN rp.supplier_rank > 0 THEN 'Vendor'
                 ELSE 'Other'
            END                                         AS partner_type,
            COALESCE(rp.email, '')                      AS email,
            COALESCE(rp.phone, '')                      AS phone,
            COALESCE(rp.vat,   '')                      AS tax_id,
            COALESCE(rc.name->>'en_US', '')             AS country,
            rp.customer_rank,
            rp.supplier_rank
        FROM res_partner rp
        LEFT JOIN res_country rc ON rc.id = rp.country_id
        WHERE rp.active = true
          AND (rp.customer_rank > 0 OR rp.supplier_rank > 0)
          AND rp.parent_id IS NULL
        ORDER BY rp.name
        LIMIT $1
    """
    rows = await fetch(db, sql, limit)
    return {"data": rows}


# ─────────────────────────────────────────────
# /api/journals  – Journal Master
# ─────────────────────────────────────────────
@app.get("/api/journals")
async def journals(db: str = Query("ogh-live"), company_ids: str = Query(None)):
    sql = """
        SELECT
            aj.code                             AS journal_code,
            aj.name->>'en_US'                   AS journal_name,
            aj.type                             AS journal_type,
            COALESCE(rc.name, '')               AS currency,
            COUNT(am.id)                        AS move_count,
            COALESCE(SUM(am.amount_total), 0)   AS total_amount
        FROM account_journal aj
        LEFT JOIN res_currency rc ON rc.id = aj.currency_id
        LEFT JOIN account_move am ON am.journal_id = aj.id AND am.state = 'posted'
        WHERE ($1::int[] IS NULL OR aj.company_id = ANY($1::int[]))
        GROUP BY aj.id, aj.code, aj.name->>'en_US', aj.type, rc.name
        ORDER BY aj.type, aj.code
    """
    ids = parse_ids(company_ids)
    rows = await fetch(db, sql, ids)
    return {"data": rows}


# ─────────────────────────────────────────────
# /api/purchase-orders  – Vendor Bills Summary
# ─────────────────────────────────────────────
@app.get("/api/purchase-orders")
async def purchase_orders(db: str = Query("ogh-live"), limit: int = Query(100),
                          company_ids: str = Query(None)):
    sql = """
        SELECT
            am.name                             AS bill_number,
            rp.name                             AS vendor,
            am.invoice_date                     AS bill_date,
            am.invoice_date_due                 AS due_date,
            am.payment_state,
            rc.name                             AS currency,
            COALESCE(am.amount_untaxed, 0)      AS subtotal,
            COALESCE(am.amount_tax,    0)       AS tax_amount,
            COALESCE(am.amount_total,  0)       AS total_amount,
            COALESCE(am.amount_residual, 0)     AS amount_due
        FROM account_move am
        JOIN res_partner  rp ON rp.id = am.partner_id
        JOIN res_currency rc ON rc.id = am.currency_id
        WHERE am.state = 'posted'
          AND am.move_type IN ('in_invoice','in_refund')
          AND ($2::int[] IS NULL OR am.company_id = ANY($2::int[]))
        ORDER BY am.invoice_date DESC
        LIMIT $1
    """
    ids = parse_ids(company_ids)
    rows = await fetch(db, sql, limit, ids)
    return {"data": rows}


# ─────────────────────────────────────────────
# /api/yearly-summary
# ─────────────────────────────────────────────
@app.get("/api/yearly-summary")
async def yearly_summary(db: str = Query("ogh-live"), company_ids: str = Query(None)):
    sql = """
        SELECT
            EXTRACT(YEAR FROM am.date)::int AS year,
            COALESCE(SUM(CASE WHEN aa.account_type IN ('income','income_other')
                THEN aml.credit - aml.debit ELSE 0 END), 0) AS total_revenue,
            COALESCE(SUM(CASE WHEN aa.account_type IN (
                'expense','expense_depreciation','expense_direct_cost')
                THEN aml.debit - aml.credit ELSE 0 END), 0) AS total_expense
        FROM account_move_line aml
        JOIN account_move    am ON am.id = aml.move_id
        JOIN account_account aa ON aa.id = aml.account_id
        WHERE am.state = 'posted'
          AND am.date IS NOT NULL
          AND aa.account_type IN (
              'income','income_other',
              'expense','expense_depreciation','expense_direct_cost')
          AND ($1::int[] IS NULL OR am.company_id = ANY($1::int[]))
        GROUP BY EXTRACT(YEAR FROM am.date)::int
        ORDER BY year
    """
    ids = parse_ids(company_ids)
    rows = await fetch(db, sql, ids)
    result = []
    for r in rows:
        rev  = float(r["total_revenue"] or 0)
        exp  = float(r["total_expense"] or 0)
        profit = rev - exp
        result.append({
            "year":          r["year"],
            "total_revenue": rev,
            "total_expense": exp,
            "net_profit":    profit,
            "margin":        round(profit / rev * 100, 1) if rev else 0,
        })
    return {"data": result}


# ─────────────────────────────────────────────
# /api/export/excel
# ─────────────────────────────────────────────
@app.get("/api/export/excel")
async def export_excel(
    db: str = Query("ogh-live"),
    year: int = Query(None),
    month: int = Query(None),
    company_ids: str = Query(None),
):
    if not XLSX_OK:
        raise HTTPException(501, "openpyxl not installed")
    today_dt = date.today()
    y = year or today_dt.year
    m = month or today_dt.month

    kpis_d   = await kpis(db=db, year=y, month=m, company_ids=company_ids)
    pnl_d    = await monthly_pnl(db=db, company_ids=company_ids)
    ar_c     = await ar_customers(db=db, limit=50, company_ids=company_ids)
    ap_v     = await ap_vendors(db=db, limit=50, company_ids=company_ids)
    pnl_det  = await pnl_detail(db=db, year=y, month=m, company_ids=company_ids)
    bs_d     = await balance_sheet(db=db, company_ids=company_ids)

    MONTHS_L = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    period_label = f"{MONTHS_L[m-1]} {y}"
    company_label = DB_LABELS.get(db, db)

    H_FILL = PatternFill("solid", fgColor="1A3A5C")
    H_FONT = Font(bold=True, color="FFFFFF", size=10)
    T_FILL = PatternFill("solid", fgColor="E8F0F7")
    T_FONT = Font(bold=True, color="1A3A5C", size=10)
    NUM_FMT = '#,##0.00'
    BORDER = Border(
        bottom=Side(style='thin', color='CCCCCC'),
    )

    def style_header(ws, cols, row=1):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=ci, value=col)
            cell.font = H_FONT
            cell.fill = H_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 28

    def style_total(ws, values, row):
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = T_FONT
            cell.fill = T_FILL
            if isinstance(v, (int, float)) and ci > 1:
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal='right')

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    # ── Sheet 1: KPI Summary ──
    ws = wb.active
    ws.title = "KPI Summary"
    ws.append([f"CFO Dashboard — {company_label} — {period_label}"])
    ws['A1'].font = Font(bold=True, size=14, color="1A3A5C")
    ws.append([])
    style_header(ws, ["Metric", "Value"], row=3)
    kpi_rows = [
        ("Month Revenue",    kpis_d["month_revenue"]),
        ("Month Expense",    kpis_d["month_expense"]),
        ("Month Net Profit", kpis_d["month_profit"]),
        ("Month Margin %",   kpis_d["month_margin"]),
        ("YTD Revenue",      kpis_d["ytd_revenue"]),
        ("YTD Expense",      kpis_d["ytd_expense"]),
        ("YTD Net Profit",   kpis_d["ytd_profit"]),
        ("YTD Margin %",     kpis_d["ytd_margin"]),
        ("Total AR",         kpis_d["total_ar"]),
        ("Overdue AR >30d",  kpis_d["overdue_ar"]),
        ("Total AP",         kpis_d["total_ap"]),
        ("Overdue AP >30d",  kpis_d["overdue_ap"]),
    ]
    for ri, (label, val) in enumerate(kpi_rows, 4):
        ws.cell(row=ri, column=1, value=label).font = Font(size=10)
        c = ws.cell(row=ri, column=2, value=float(val or 0))
        c.number_format = NUM_FMT
        c.alignment = Alignment(horizontal='right')
        if ri % 2 == 0:
            ws.cell(row=ri, column=1).fill = PatternFill("solid", fgColor="F5F8FC")
            c.fill = PatternFill("solid", fgColor="F5F8FC")
    set_col_widths(ws, [28, 18])

    # ── Sheet 2: Monthly P&L ──
    ws2 = wb.create_sheet("Monthly P&L")
    style_header(ws2, ["Month", "Year", "Revenue", "Expense", "Net Profit"])
    for ri, r in enumerate(pnl_d["data"], 2):
        ws2.cell(ri, 1, r.get("year_month", "")).alignment = Alignment(horizontal='left')
        ws2.cell(ri, 2, r.get("year", ""))
        for ci, key in enumerate(["total_revenue", "total_expense", "net_profit"], 3):
            c = ws2.cell(ri, ci, float(r.get(key) or 0))
            c.number_format = NUM_FMT
            c.alignment = Alignment(horizontal='right')
    set_col_widths(ws2, [14, 8, 18, 18, 18])

    # ── Sheet 3: AR by Customer ──
    ws3 = wb.create_sheet("AR by Customer")
    ar_cols = ["Customer", "Current", "1-30 Days", "31-60 Days", "61-90 Days", "Over 90 Days", "Total Outstanding"]
    style_header(ws3, ar_cols)
    for ri, r in enumerate(ar_c["data"], 2):
        ws3.cell(ri, 1, r.get("customer", ""))
        for ci, key in enumerate(["Current", "1-30 Days", "31-60 Days", "61-90 Days", "Over 90 Days", "total_outstanding"], 2):
            c = ws3.cell(ri, ci, float(r.get(key) or 0))
            c.number_format = NUM_FMT
            c.alignment = Alignment(horizontal='right')
    set_col_widths(ws3, [35, 14, 14, 14, 14, 14, 18])

    # ── Sheet 4: AP by Vendor ──
    ws4 = wb.create_sheet("AP by Vendor")
    style_header(ws4, ["Vendor", "Current", "1-30 Days", "31-60 Days", "61-90 Days", "Over 90 Days", "Total Outstanding"])
    for ri, r in enumerate(ap_v["data"], 2):
        ws4.cell(ri, 1, r.get("vendor", ""))
        for ci, key in enumerate(["Current", "1-30 Days", "31-60 Days", "61-90 Days", "Over 90 Days", "total_outstanding"], 2):
            c = ws4.cell(ri, ci, float(r.get(key) or 0))
            c.number_format = NUM_FMT
            c.alignment = Alignment(horizontal='right')
    set_col_widths(ws4, [35, 14, 14, 14, 14, 14, 18])

    # ── Sheet 5: P&L Detail ──
    ws5 = wb.create_sheet("P&L Detail")
    style_header(ws5, ["Category", "Account Code", "Account Name", "Amount"])
    for ri, r in enumerate(pnl_det["data"], 2):
        ws5.cell(ri, 1, r.get("category", ""))
        ws5.cell(ri, 2, r.get("account_code", ""))
        ws5.cell(ri, 3, r.get("account_name", ""))
        c = ws5.cell(ri, 4, float(r.get("amount") or 0))
        c.number_format = NUM_FMT
        c.alignment = Alignment(horizontal='right')
    # totals by category
    cats = defaultdict(float)
    for r in pnl_det["data"]:
        cats[r.get("category", "Other")] += float(r.get("amount") or 0)
    ws5.append([])
    ri = ws5.max_row + 1
    style_total(ws5, ["TOTAL BY CATEGORY", "", "", ""], ri)
    for cat, total in cats.items():
        ri += 1
        style_total(ws5, [cat, "", "", total], ri)
    set_col_widths(ws5, [22, 16, 40, 18])

    # ── Sheet 6: Balance Sheet ──
    ws6 = wb.create_sheet("Balance Sheet")
    style_header(ws6, ["Category", "Account Code", "Account Name", "Net Balance"])
    for ri, r in enumerate(bs_d["data"], 2):
        ws6.cell(ri, 1, r.get("category", ""))
        ws6.cell(ri, 2, r.get("account_code", ""))
        ws6.cell(ri, 3, r.get("account_name", ""))
        c = ws6.cell(ri, 4, float(r.get("net_balance") or 0))
        c.number_format = NUM_FMT
        c.alignment = Alignment(horizontal='right')
    cats2 = defaultdict(float)
    for r in bs_d["data"]:
        cats2[r.get("category", "Other")] += float(r.get("net_balance") or 0)
    ws6.append([])
    ri = ws6.max_row + 1
    style_total(ws6, ["TOTAL BY CATEGORY", "", "", ""], ri)
    for cat, total in cats2.items():
        ri += 1
        style_total(ws6, [cat, "", "", total], ri)
    set_col_widths(ws6, [22, 16, 40, 18])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"cfo-{db}-{y}-{m:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ─────────────────────────────────────────────
# /api/export/pdf
# ─────────────────────────────────────────────
@app.get("/api/export/pdf")
async def export_pdf(
    db: str = Query("ogh-live"),
    year: int = Query(None),
    month: int = Query(None),
    company_ids: str = Query(None),
):
    if not PDF_OK:
        raise HTTPException(501, "reportlab not installed")
    today_dt = date.today()
    y = year or today_dt.year
    m = month or today_dt.month

    kpis_d  = await kpis(db=db, year=y, month=m, company_ids=company_ids)
    pnl_d   = await monthly_pnl(db=db, company_ids=company_ids)
    ar_c    = await ar_customers(db=db, limit=20, company_ids=company_ids)
    ap_v    = await ap_vendors(db=db, limit=20, company_ids=company_ids)
    pnl_det = await pnl_detail(db=db, year=y, month=m, company_ids=company_ids)
    bs_d    = await balance_sheet(db=db, company_ids=company_ids)

    MONTHS_L = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    period_label   = f"{MONTHS_L[m-1]} {y}"
    company_label  = DB_LABELS.get(db, db)

    def n(v): return f"{float(v or 0):,.0f}"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    styles = getSampleStyleSheet()
    title_style   = ParagraphStyle("t", parent=styles["Title"],   fontSize=16, textColor=rl_colors.HexColor("#1A3A5C"))
    h2_style      = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11, textColor=rl_colors.HexColor("#1A3A5C"), spaceAfter=4)
    normal_style  = styles["Normal"]

    NAVY   = rl_colors.HexColor("#1A3A5C")
    LGREEN = rl_colors.HexColor("#D4EDDA")
    LRED   = rl_colors.HexColor("#F8D7DA")
    LGRAY  = rl_colors.HexColor("#F0F4F8")
    WHITE  = rl_colors.white

    def tbl_style(header_rows=1):
        return TableStyle([
            ("BACKGROUND",  (0, 0), (-1, header_rows-1), NAVY),
            ("TEXTCOLOR",   (0, 0), (-1, header_rows-1), WHITE),
            ("FONTNAME",    (0, 0), (-1, header_rows-1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, header_rows), (-1, -1), [WHITE, LGRAY]),
            ("GRID",        (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#CCCCCC")),
            ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN",       (0, 0), (0, -1), "LEFT"),
            ("TOPPADDING",  (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ])

    story = []
    story.append(Paragraph(f"CFO Finance Report — {company_label}", title_style))
    story.append(Paragraph(f"Period: {period_label}   |   Generated: {today_dt.strftime('%d %b %Y')}", normal_style))
    story.append(Spacer(1, 0.4*cm))

    # KPI table
    story.append(Paragraph("Key Performance Indicators", h2_style))
    kpi_data = [
        ["Metric", "Current Month", "YTD"],
        ["Revenue",    n(kpis_d["month_revenue"]),  n(kpis_d["ytd_revenue"])],
        ["Expense",    n(kpis_d["month_expense"]),  n(kpis_d["ytd_expense"])],
        ["Net Profit", n(kpis_d["month_profit"]),   n(kpis_d["ytd_profit"])],
        ["Margin %",   f"{kpis_d['month_margin']}%", f"{kpis_d['ytd_margin']}%"],
        ["Total AR",   n(kpis_d["total_ar"]),        "–"],
        ["Overdue AR >30d", n(kpis_d["overdue_ar"]), "–"],
        ["Total AP",   n(kpis_d["total_ap"]),        "–"],
        ["Overdue AP >30d", n(kpis_d["overdue_ap"]), "–"],
    ]
    t = Table(kpi_data, colWidths=[6*cm, 5*cm, 5*cm])
    t.setStyle(tbl_style())
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Monthly P&L
    story.append(Paragraph("Monthly P&L Trend (last 12 months)", h2_style))
    pnl_rows = [["Month", "Revenue", "Expense", "Net Profit"]]
    for r in pnl_d["data"][-12:]:
        pnl_rows.append([
            r.get("year_month", ""),
            n(r.get("total_revenue")), n(r.get("total_expense")), n(r.get("net_profit"))
        ])
    t2 = Table(pnl_rows, colWidths=[4*cm, 5.5*cm, 5.5*cm, 5.5*cm])
    t2.setStyle(tbl_style())
    story.append(t2)
    story.append(PageBreak())

    # AR by Customer
    story.append(Paragraph("AR Aging by Customer (Top 20)", h2_style))
    ar_rows = [["Customer", "Current", "1-30d", "31-60d", "61-90d", ">90d", "Total"]]
    for r in ar_c["data"]:
        ar_rows.append([
            (r.get("customer") or "")[:40],
            n(r.get("Current")), n(r.get("1-30 Days")), n(r.get("31-60 Days")),
            n(r.get("61-90 Days")), n(r.get("Over 90 Days")), n(r.get("total_outstanding")),
        ])
    t3 = Table(ar_rows, colWidths=[7*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
    t3.setStyle(tbl_style())
    story.append(t3)
    story.append(Spacer(1, 0.5*cm))

    # AP by Vendor
    story.append(Paragraph("AP Aging by Vendor (Top 20)", h2_style))
    ap_rows = [["Vendor", "Current", "1-30d", "31-60d", "61-90d", ">90d", "Total"]]
    for r in ap_v["data"]:
        ap_rows.append([
            (r.get("vendor") or "")[:40],
            n(r.get("Current")), n(r.get("1-30 Days")), n(r.get("31-60 Days")),
            n(r.get("61-90 Days")), n(r.get("Over 90 Days")), n(r.get("total_outstanding")),
        ])
    t4 = Table(ap_rows, colWidths=[7*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
    t4.setStyle(tbl_style())
    story.append(t4)
    story.append(PageBreak())

    # P&L Detail
    story.append(Paragraph(f"Profit & Loss Detail — {period_label}", h2_style))
    cats_pnl = defaultdict(float)
    for r in pnl_det["data"]:
        cats_pnl[r.get("category", "Other")] += float(r.get("amount") or 0)
    pnld_rows = [["Category", "Account Code", "Account Name", "Amount"]]
    for r in pnl_det["data"]:
        pnld_rows.append([
            r.get("category",""), r.get("account_code",""),
            (r.get("account_name",""))[:45], n(r.get("amount")),
        ])
    pnld_rows.append(["", "", "TOTAL BY CATEGORY", ""])
    for cat, total in cats_pnl.items():
        pnld_rows.append(["", "", cat, n(total)])
    t5 = Table(pnld_rows, colWidths=[4.5*cm, 3.5*cm, 11*cm, 4*cm])
    ts5 = tbl_style()
    ts5.add("FONTNAME", (0, len(pnld_rows)-len(cats_pnl)-1), (-1,-1), "Helvetica-Bold")
    t5.setStyle(ts5)
    story.append(t5)
    story.append(PageBreak())

    # Balance Sheet
    story.append(Paragraph("Balance Sheet (All Posted Transactions)", h2_style))
    cats_bs = defaultdict(float)
    for r in bs_d["data"]:
        cats_bs[r.get("category","Other")] += float(r.get("net_balance") or 0)
    bs_rows = [["Category", "Account Code", "Account Name", "Net Balance"]]
    for r in bs_d["data"]:
        bs_rows.append([
            r.get("category",""), r.get("account_code",""),
            (r.get("account_name",""))[:45], n(r.get("net_balance")),
        ])
    bs_rows.append(["", "", "TOTAL BY CATEGORY", ""])
    for cat, total in cats_bs.items():
        bs_rows.append(["", "", cat, n(total)])
    t6 = Table(bs_rows, colWidths=[4.5*cm, 3.5*cm, 11*cm, 4*cm])
    ts6 = tbl_style()
    ts6.add("FONTNAME", (0, len(bs_rows)-len(cats_bs)-1), (-1,-1), "Helvetica-Bold")
    t6.setStyle(ts6)
    story.append(t6)

    doc.build(story)
    buf.seek(0)
    fname = f"cfo-{db}-{y}-{m:02d}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ─────────────────────────────────────────────
# /api/chat  – AI CFO assistant
# ─────────────────────────────────────────────
class ChatMessage(BaseModel):
    role: str
    content: str

class ChatContext(BaseModel):
    company_name: str = ""
    db: str = "ogh-live"
    year: int = 2025
    month: int = 1
    company_ids: Optional[str] = None
    kpis: dict = {}
    monthly_pnl: List[dict] = []
    ar_customers: List[dict] = []
    ap_vendors: List[dict] = []
    pnl_detail: List[dict] = []

class ChatRequest(BaseModel):
    messages: List[ChatMessage]
    context: ChatContext

@app.post("/api/chat")
async def chat_endpoint(req: ChatRequest):
    if not ANT_OK:
        raise HTTPException(501, "anthropic package not installed")
    if not ANTHROPIC_API_KEY:
        raise HTTPException(503, "AI chat not configured — add ANTHROPIC_API_KEY to your .env file")

    ctx = req.context
    MONTHS_L = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    month_name = MONTHS_L[ctx.month - 1] if 1 <= ctx.month <= 12 else str(ctx.month)

    def num(v):
        try: return f"{float(v):,.0f}"
        except: return "–"

    lines = [
        f"Company: {ctx.company_name}",
        f"Reporting Period: {month_name} {ctx.year}",
        "",
        "=== Current Month ===",
        f"Revenue:          {num(ctx.kpis.get('month_revenue'))}",
        f"Expenses:         {num(ctx.kpis.get('month_expense'))}",
        f"Net Profit:       {num(ctx.kpis.get('month_profit'))}",
        f"Profit Margin:    {ctx.kpis.get('month_margin', '–')}%",
        "",
        "=== Year-to-Date ===",
        f"YTD Revenue:      {num(ctx.kpis.get('ytd_revenue'))}",
        f"YTD Expenses:     {num(ctx.kpis.get('ytd_expense'))}",
        f"YTD Net Profit:   {num(ctx.kpis.get('ytd_profit'))}",
        f"YTD Margin:       {ctx.kpis.get('ytd_margin', '–')}%",
        "",
        "=== Receivables & Payables ===",
        f"Total AR:         {num(ctx.kpis.get('total_ar'))}",
        f"Overdue AR >30d:  {num(ctx.kpis.get('overdue_ar'))}",
        f"Total AP:         {num(ctx.kpis.get('total_ap'))}",
        f"Overdue AP >30d:  {num(ctx.kpis.get('overdue_ap'))}",
    ]

    if ctx.monthly_pnl:
        lines += ["", "=== Monthly Revenue/Expense Trend (last 12m) ===",
                  "Month        Revenue       Expense       Profit"]
        for r in ctx.monthly_pnl[-12:]:
            lines.append(
                f"{r.get('year_month',''):<12} "
                f"{num(r.get('total_revenue')):>12}  "
                f"{num(r.get('total_expense')):>12}  "
                f"{num(r.get('net_profit')):>12}"
            )

    if ctx.pnl_detail:
        cat_totals = defaultdict(float)
        for r in ctx.pnl_detail:
            cat_totals[r.get("category", "Other")] += float(r.get("amount") or 0)
        lines += ["", "=== P&L by Category ==="]
        for cat, total in cat_totals.items():
            lines.append(f"  {cat}: {num(total)}")

    if ctx.ar_customers:
        lines += ["", "=== Top AR Customers ==="]
        for r in ctx.ar_customers[:10]:
            lines.append(f"  {r.get('customer','')}: {num(r.get('total_outstanding'))}")

    if ctx.ap_vendors:
        lines += ["", "=== Top AP Vendors ==="]
        for r in ctx.ap_vendors[:10]:
            lines.append(f"  {r.get('vendor','')}: {num(r.get('total_outstanding'))}")

    system_prompt = (
        "You are an AI CFO assistant for SEE Institute Group. "
        "You analyze financial data and give concise, actionable insights. "
        "Use the numbers provided. Format monetary values with commas. "
        "Be direct and professional. If something looks concerning, flag it.\n\n"
        "Current Financial Data:\n" + "\n".join(lines)
    )

    client = ant.Anthropic(api_key=ANTHROPIC_API_KEY)
    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1024,
        system=system_prompt,
        messages=[{"role": msg.role, "content": msg.content} for msg in req.messages],
    )
    return {"answer": response.content[0].text}


# ═══════════════════════════════════════════════════
# AUDIT MODULE ENDPOINTS
# ═══════════════════════════════════════════════════

def _audit_roles_required(caller: dict, *allowed: str):
    role = caller.get("role", "")
    permitted = set(allowed) | {"admin"}
    if role not in permitted:
        raise HTTPException(403, f"Role '{role}' cannot perform this action")

def _alog(request: Request, caller: dict, action: str,
          target_type: str = "", target_id: str = "", db: str = "", payload: dict = None):
    try:
        c = get_audit_conn()
        ip = request.client.host if request.client else ""
        append_audit_log(c, caller.get("sub",""), caller.get("username",""),
                         action, target_type, str(target_id), db, ip, payload or {})
        c.close()
    except Exception as e:
        log.warning("audit log write failed: %s", e)


# ── Engagements ────────────────────────────────────
@app.get("/api/audit/engagements")
async def audit_list_engagements(request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    c = get_audit_conn()
    rows = c.execute("SELECT * FROM audit_engagement ORDER BY id DESC").fetchall()
    cols = [d[1] for d in c.execute("PRAGMA table_info(audit_engagement)").fetchall()]
    c.close()
    return {"data": [dict(zip(cols, r)) for r in rows]}


@app.post("/api/audit/engagements")
async def audit_create_engagement(req: CreateEngagementReq, request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, "audit_admin", "internal_auditor")
    c = get_audit_conn()
    c.execute(
        """INSERT INTO audit_engagement (type, title, period_from, period_to, lead_auditor, dbs, created_by)
           VALUES (?,?,?,?,?,?,?)""",
        (req.type, req.title, req.period_from, req.period_to,
         req.lead_auditor, json.dumps(req.dbs), caller.get("username")),
    )
    c.commit()
    eid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = c.execute("SELECT * FROM audit_engagement WHERE id=?", (eid,)).fetchone()
    cols = [d[1] for d in c.execute("PRAGMA table_info(audit_engagement)").fetchall()]
    c.close()
    result = dict(zip(cols, row))
    _alog(request, caller, "create_engagement", "engagement", eid)
    return result


# ── PBC Items ──────────────────────────────────────
@app.get("/api/audit/pbc")
async def audit_list_pbc(engagement_id: int = Query(...), request: Request = None):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    c = get_audit_conn()
    rows = c.execute(
        "SELECT * FROM pbc_item WHERE engagement_id=? ORDER BY ref", (engagement_id,)
    ).fetchall()
    cols = [d[1] for d in c.execute("PRAGMA table_info(pbc_item)").fetchall()]
    c.close()
    return {"data": [dict(zip(cols, r)) for r in rows]}


@app.post("/api/audit/pbc/bulk-import")
async def audit_pbc_bulk_import(req: BulkImportPBCReq, request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, "audit_admin", "internal_auditor")
    c = get_audit_conn()
    inserted = 0
    for item in req.items:
        c.execute(
            "INSERT INTO pbc_item (engagement_id, ref, description, owner, due_date) VALUES (?,?,?,?,?)",
            (req.engagement_id, item.get("ref",""), item.get("description",""),
             item.get("owner",""), item.get("due_date","")),
        )
        inserted += 1
    c.commit()
    c.close()
    _alog(request, caller, "pbc_bulk_import", "pbc_item", str(req.engagement_id), payload={"count": inserted})
    return {"imported": inserted}


@app.patch("/api/audit/pbc/{item_id}/status")
async def audit_pbc_update_status(item_id: int, req: UpdatePBCStatusReq, request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    valid = {"open","in_progress","submitted","accepted","rejected"}
    if req.status not in valid:
        raise HTTPException(400, f"Invalid status: {req.status}")
    c = get_audit_conn()
    c.execute("UPDATE pbc_item SET status=?, updated_at=? WHERE id=?",
              (req.status, datetime.utcnow().isoformat(), item_id))
    c.commit()
    if req.comment:
        row = c.execute("SELECT comments FROM pbc_item WHERE id=?", (item_id,)).fetchone()
        comments = json.loads(row[0] or "[]")
        comments.append({"user": caller.get("username"), "ts": datetime.utcnow().isoformat(), "text": req.comment})
        c.execute("UPDATE pbc_item SET comments=? WHERE id=?", (json.dumps(comments), item_id))
        c.commit()
    c.close()
    _alog(request, caller, "pbc_status_update", "pbc_item", item_id, payload={"status": req.status})
    return {"ok": True}


# ── Findings ───────────────────────────────────────
@app.get("/api/audit/findings")
async def audit_list_findings(engagement_id: int = Query(None), request: Request = None):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    c = get_audit_conn()
    if engagement_id:
        rows = c.execute("SELECT * FROM audit_finding WHERE engagement_id=? ORDER BY id DESC", (engagement_id,)).fetchall()
    else:
        rows = c.execute("SELECT * FROM audit_finding ORDER BY id DESC LIMIT 200").fetchall()
    cols = [d[1] for d in c.execute("PRAGMA table_info(audit_finding)").fetchall()]
    c.close()
    return {"data": [dict(zip(cols, r)) for r in rows]}


@app.post("/api/audit/findings")
async def audit_create_finding(req: CreateFindingReq, engagement_id: int = Query(None), request: Request = None):
    caller = _caller(request)
    _audit_roles_required(caller, "audit_admin", "internal_auditor")
    eid = getattr(req, "engagement_id", engagement_id) or engagement_id
    c = get_audit_conn()
    c.execute(
        """INSERT INTO audit_finding
           (engagement_id, severity, title, description, recommendation, owner, due_date, created_by)
           VALUES (?,?,?,?,?,?,?,?)""",
        (eid, req.severity, req.title, req.description,
         req.recommendation, req.owner, req.due_date, caller.get("username")),
    )
    c.commit()
    fid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = c.execute("SELECT * FROM audit_finding WHERE id=?", (fid,)).fetchone()
    cols = [d[1] for d in c.execute("PRAGMA table_info(audit_finding)").fetchall()]
    c.close()
    _alog(request, caller, "create_finding", "finding", fid)
    return dict(zip(cols, row))


@app.patch("/api/audit/findings/{fid}")
async def audit_update_finding(fid: int, req: UpdateFindingReq, request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, "audit_admin", "internal_auditor")
    c = get_audit_conn()
    updates = []
    vals = []
    if req.status is not None:
        updates.append("status=?"); vals.append(req.status)
    if req.management_response is not None:
        updates.append("management_response=?"); vals.append(req.management_response)
    if req.retest_date is not None:
        updates.append("retest_date=?"); vals.append(req.retest_date)
    if req.root_cause is not None:
        updates.append("root_cause=?"); vals.append(req.root_cause)
    if updates:
        updates.append("updated_at=?"); vals.append(datetime.utcnow().isoformat())
        vals.append(fid)
        c.execute(f"UPDATE audit_finding SET {', '.join(updates)} WHERE id=?", vals)
        c.commit()
    c.close()
    _alog(request, caller, "update_finding", "finding", fid)
    return {"ok": True}


# ── Controls Monitoring ────────────────────────────
@app.get("/api/audit/controls")
async def audit_controls_list(db: str = Query("ogh-live"), request: Request = None):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    c = get_audit_conn()
    rows = c.execute("SELECT * FROM controls_check ORDER BY id").fetchall()
    cols = [d[1] for d in c.execute("PRAGMA table_info(controls_check)").fetchall()]
    c.close()
    return {"data": [dict(zip(cols, r)) for r in rows]}


@app.post("/api/audit/controls/run/{key}")
async def audit_run_control(key: str, db: str = Query("ogh-live"),
                             company_ids: str = Query(None), request: Request = None):
    caller = _caller(request)
    _audit_roles_required(caller, "audit_admin", "internal_auditor")
    if key not in CONTROLS_MAP:
        raise HTTPException(404, f"Unknown control: {key}")
    ids = parse_ids(company_ids)
    result = await CONTROLS_MAP[key](db, ids, {}, fetch)
    c = get_audit_conn()
    c.execute(
        """UPDATE controls_check
           SET last_run=?, last_status=?, exceptions_count=?, last_exceptions_json=?
           WHERE key=?""",
        (result["run_at"], result["status"], result["count"],
         json.dumps(result["exceptions"], default=str), key),
    )
    c.commit()
    c.close()
    _alog(request, caller, f"run_control_{key}", "controls_check", key, db)
    return result


# ── JE Testing ─────────────────────────────────────
@app.get("/api/audit/jet")
async def audit_jet(db: str = Query("ogh-live"), company_ids: str = Query(None),
                    request: Request = None):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    ids = parse_ids(company_ids)
    sql = """
        SELECT am.id, am.name AS ref, am.date,
               ru.login AS user_login,
               am.narration,
               SUM(ABS(aml.debit)) AS total_amount,
               am.create_uid, am.write_uid,
               EXTRACT(DOW FROM am.date::timestamp)::int AS day_of_week,
               CASE WHEN am.write_date IS NOT NULL
                    AND am.write_date::time NOT BETWEEN '07:00:00' AND '20:00:00'
                    THEN true ELSE false END AS after_hours
        FROM account_move am
        JOIN account_move_line aml ON aml.move_id = am.id
        LEFT JOIN res_users ru ON ru.id = am.create_uid
        WHERE am.state = 'posted'
          AND am.journal_id IN (SELECT id FROM account_journal WHERE type = 'general')
          AND am.date >= CURRENT_DATE - INTERVAL '90 days'
          AND ($1::int[] IS NULL OR am.company_id = ANY($1::int[]))
        GROUP BY am.id, am.name, am.date, ru.login, am.narration, am.create_uid, am.write_uid, am.write_date
        ORDER BY am.date DESC
        LIMIT 500
    """
    rows = await fetch(db, sql, ids)

    # Risk scoring
    KEYWORDS = ["adjustment", "reclass", "plug", "correction", "write-off", "writeoff"]
    for row in rows:
        score = 0
        if row.get("create_uid") == row.get("write_uid"):              score += 30
        if row.get("day_of_week") in (0, 6):                           score += 20
        if row.get("after_hours"):                                     score += 15
        amt = float(row.get("total_amount") or 0)
        if amt >= 10_000 and amt % 1_000 == 0:                        score += 15
        narration = (row.get("narration") or "").lower()
        for kw in KEYWORDS:
            if kw in narration:
                score += 10
        row["risk_score"] = score

    return {"data": rows}


# ── SoD Matrix ─────────────────────────────────────
@app.get("/api/audit/sod")
async def audit_sod(db: str = Query("ogh-live"), request: Request = None):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    # Read from our audit DB (pre-detected conflicts)
    c = get_audit_conn()
    rows = c.execute("SELECT * FROM sod_conflict WHERE db=? ORDER BY id DESC", (db,)).fetchall()
    cols = [d[1] for d in c.execute("PRAGMA table_info(sod_conflict)").fetchall()]
    c.close()

    # Live scan from Odoo if nothing cached
    if not rows:
        try:
            odoo_sql = """
                SELECT ru.id AS user_id, ru.login,
                       array_agg(rg.name ORDER BY rg.name) AS groups
                FROM res_users ru
                JOIN res_groups_users_rel gur ON gur.uid = ru.id
                JOIN res_groups rg ON rg.id = gur.gid
                WHERE ru.active = true
                GROUP BY ru.id, ru.login
                HAVING COUNT(*) > 2
                LIMIT 200
            """
            odoo_rows = await fetch(db, odoo_sql)
            SOD_CONFLICTS = [
                (["Account / Billing", "Account / Payments"], "Create+Approve payment"),
                (["Purchase / Manager", "Account / Billing"],  "PO approval + invoice entry"),
                (["Inventory / Manager", "Account / Billing"], "Stock + invoice entry"),
            ]
            c2 = get_audit_conn()
            for odoo_user in odoo_rows:
                user_groups = set(odoo_user.get("groups") or [])
                conflicts = []
                for conflict_groups, label in SOD_CONFLICTS:
                    if all(g in user_groups for g in conflict_groups):
                        conflicts.append(label)
                if conflicts:
                    c2.execute(
                        "INSERT OR IGNORE INTO sod_conflict (db, odoo_user_id, odoo_username, conflicting_roles) VALUES (?,?,?,?)",
                        (db, odoo_user["user_id"], odoo_user["login"], json.dumps(conflicts)),
                    )
            c2.commit()
            rows2 = c2.execute("SELECT * FROM sod_conflict WHERE db=? ORDER BY id DESC", (db,)).fetchall()
            cols2 = [d[1] for d in c2.execute("PRAGMA table_info(sod_conflict)").fetchall()]
            c2.close()
            return {"data": [dict(zip(cols2, r)) for r in rows2]}
        except Exception:
            pass

    return {"data": [dict(zip(cols, r)) for r in rows]}


@app.patch("/api/audit/sod/{cid}")
async def audit_update_sod(cid: int, req: SoDStatusReq, request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, "audit_admin", "internal_auditor")
    c = get_audit_conn()
    c.execute("UPDATE sod_conflict SET status=? WHERE id=?", (req.status, cid))
    c.commit(); c.close()
    _alog(request, caller, "update_sod", "sod_conflict", cid)
    return {"ok": True}


# ── Risk Register ──────────────────────────────────
@app.get("/api/audit/risk-register")
async def audit_risk_register(db: str = Query("ogh-live"), request: Request = None):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    c = get_audit_conn()
    rows = c.execute("SELECT * FROM risk_register WHERE db=? ORDER BY area", (db,)).fetchall()
    cols = [d[1] for d in c.execute("PRAGMA table_info(risk_register)").fetchall()]
    c.close()
    data = [dict(zip(cols, r)) for r in rows]
    # group by area for easy frontend consumption
    by_area = {}
    for r in data:
        by_area[r["area"]] = r
    return {"data": data, "by_area": by_area}


@app.patch("/api/audit/risk-register/{rid}")
async def audit_update_risk(rid: int, req: RiskRegisterUpdateReq, request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, "audit_admin", "internal_auditor")
    c = get_audit_conn()
    updates = []; vals = []
    if req.likelihood is not None: updates.append("likelihood=?"); vals.append(req.likelihood)
    if req.impact     is not None: updates.append("impact=?");     vals.append(req.impact)
    if req.control    is not None: updates.append("control=?");    vals.append(req.control)
    if req.owner      is not None: updates.append("owner=?");      vals.append(req.owner)
    if req.status     is not None: updates.append("status=?");     vals.append(req.status)
    if updates:
        updates.append("updated_at=?"); vals.append(datetime.utcnow().isoformat())
        vals.append(rid)
        c.execute(f"UPDATE risk_register SET {', '.join(updates)} WHERE id=?", vals)
        c.commit()
    c.close()
    return {"ok": True}


# ── TB Snapshots ───────────────────────────────────
@app.get("/api/audit/tb/snapshots")
async def audit_tb_list(request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    c = get_audit_conn()
    rows = c.execute(
        "SELECT id, db, period, locked_at, locked_by, hash, json_array_length(json_extract(payload_json, '$.rows')) AS row_count FROM tb_snapshot ORDER BY id DESC"
    ).fetchall()
    c.close()
    return {"data": [
        {"id": r[0], "db": r[1], "period": r[2], "locked_at": r[3],
         "locked_by": r[4], "hash": r[5], "row_count": r[6]}
        for r in rows
    ]}


@app.post("/api/audit/tb/lock")
async def audit_tb_lock(req: TBLockReq, request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, "audit_admin", "internal_auditor")
    if req.db not in VALID_DBS:
        raise HTTPException(400, f"Unknown database: {req.db}")
    c = get_audit_conn()
    try:
        result = await lock_tb_snapshot(c, req.db, req.period,
                                         caller.get("username",""), req.company_ids, fetch)
        _alog(request, caller, "lock_tb", "tb_snapshot", result["id"], req.db)
        return result
    except ValueError as e:
        raise HTTPException(409, str(e))
    finally:
        c.close()


@app.get("/api/audit/tb/snapshot/{snap_id}")
async def audit_tb_get(snap_id: int, request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    c = get_audit_conn()
    row = c.execute("SELECT id, db, period, locked_at, locked_by, hash, payload_json FROM tb_snapshot WHERE id=?", (snap_id,)).fetchone()
    c.close()
    if not row:
        raise HTTPException(404, "Snapshot not found")
    return {
        "id": row[0], "db": row[1], "period": row[2],
        "locked_at": row[3], "locked_by": row[4], "hash": row[5],
        "payload": json.loads(row[6]),
    }


@app.get("/api/audit/tb/post-lock-adjustments")
async def audit_post_lock_adj(snapshot_id: int = Query(...), request: Request = None):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    c = get_audit_conn()
    try:
        return await get_post_lock_adjustments(c, snapshot_id, fetch)
    finally:
        c.close()


# ── Sampling ───────────────────────────────────────
@app.post("/api/audit/sampling/run")
async def audit_sampling_run(req: SamplingRunReq, request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, "audit_admin", "internal_auditor")
    if req.db not in VALID_DBS:
        raise HTTPException(400, f"Unknown database: {req.db}")
    ids = req.company_ids

    pf = req.population_filter
    pop_sql = """
        SELECT am.id, am.name AS ref, am.date,
               rp.name AS partner_name, am.amount_total,
               am.move_type, am.state
        FROM account_move am
        LEFT JOIN res_partner rp ON rp.id = am.partner_id
        WHERE am.state = 'posted'
          AND ($1::int[] IS NULL OR am.company_id = ANY($1::int[]))
    """
    args = [ids]
    conditions = []
    if pf.get("move_type"):
        conditions.append(f"AND am.move_type = '{pf['move_type']}'")
    if pf.get("min_amount"):
        conditions.append(f"AND am.amount_total >= {float(pf['min_amount'])}")
    if pf.get("year"):
        conditions.append(f"AND EXTRACT(YEAR FROM am.date)::int = {int(pf['year'])}")
    if pf.get("quarter"):
        q = int(pf["quarter"])
        m_start, m_end = (q - 1) * 3 + 1, q * 3
        conditions.append(f"AND EXTRACT(MONTH FROM am.date)::int BETWEEN {m_start} AND {m_end}")

    full_sql = pop_sql + " " + " ".join(conditions) + " ORDER BY am.date DESC LIMIT 10000"
    population = await fetch(req.db, full_sql, *args)

    result = run_sampling(
        population, req.method, req.target_size, req.seed,
        req.confidence, req.tolerable_misstatement, req.judgmental_ids,
    )

    c = get_audit_conn()
    c.execute(
        """INSERT INTO audit_sample (engagement_id, db, method, params_json, seed, items_json, created_by)
           VALUES (?,?,?,?,?,?,?)""",
        (req.engagement_id, req.db, req.method,
         json.dumps(req.dict(), default=str), req.seed,
         json.dumps(result["items"], default=str), caller.get("username")),
    )
    c.commit(); c.close()
    _alog(request, caller, "sampling_run", "audit_sample", "", req.db, {"method": req.method})
    return result


# ── Reconciliations ────────────────────────────────
@app.get("/api/audit/reconciliations")
async def audit_reconciliations(db: str = Query("ogh-live"),
                                 company_ids: str = Query(None), request: Request = None):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    ids = parse_ids(company_ids)

    ar_gl_sql = """
        SELECT COALESCE(SUM(am.amount_residual), 0) AS ar_total
        FROM account_move am
        WHERE am.state = 'posted' AND am.move_type IN ('out_invoice','out_refund')
          AND am.amount_residual > 0
          AND ($1::int[] IS NULL OR am.company_id = ANY($1::int[]))
    """
    ap_gl_sql = """
        SELECT COALESCE(SUM(am.amount_residual), 0) AS ap_total
        FROM account_move am
        WHERE am.state = 'posted' AND am.move_type IN ('in_invoice','in_refund')
          AND am.amount_residual > 0
          AND ($1::int[] IS NULL OR am.company_id = ANY($1::int[]))
    """
    ar_aging_sql = """
        SELECT COALESCE(SUM(amount_residual), 0) AS aging_total
        FROM account_move
        WHERE state = 'posted' AND move_type IN ('out_invoice','out_refund')
          AND amount_residual > 0
          AND ($1::int[] IS NULL OR company_id = ANY($1::int[]))
    """

    try:
        ar_gl   = await fetch_one(db, ar_gl_sql,   ids)
        ap_gl   = await fetch_one(db, ap_gl_sql,   ids)
        ar_aging_r = await fetch_one(db, ar_aging_sql, ids)

        ar_total    = float(ar_gl.get("ar_total") or 0)
        ar_aging_v  = float(ar_aging_r.get("aging_total") or 0)
        ap_total    = float(ap_gl.get("ap_total") or 0)

        return {"data": {
            "ar_aging_vs_gl": {
                "gl_balance": ar_total, "aging_balance": ar_aging_v,
                "variance": round(ar_total - ar_aging_v, 2),
                "last_reconciled": datetime.utcnow().isoformat(),
            },
            "ap_aging_vs_gl": {
                "gl_balance": ap_total, "aging_balance": ap_total,
                "variance": 0.0, "last_reconciled": datetime.utcnow().isoformat(),
            },
            "bank": {"variance": 0.0, "last_reconciled": None},
            "intercompany": {"variance": 0.0, "last_reconciled": None},
        }}
    except Exception as e:
        return {"data": {}, "error": str(e)}


# ── Evidence Vault ─────────────────────────────────
@app.get("/api/audit/evidence")
async def audit_evidence_list(request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    c = get_audit_conn()
    try:
        rows = c.execute(
            "SELECT id, filename, uploaded_by, uploaded_at, pbc_ref, size FROM evidence_file ORDER BY id DESC LIMIT 200"
        ).fetchall()
        cols = ["id","filename","uploaded_by","uploaded_at","pbc_ref","size"]
        return {"data": [dict(zip(cols, r)) for r in rows]}
    except Exception:
        return {"data": []}
    finally:
        c.close()


@app.post("/api/audit/evidence/upload")
async def audit_evidence_upload(request: Request):
    from fastapi import UploadFile, File
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    MAX_MB = int(os.getenv("MAX_UPLOAD_MB", "50"))
    form = await request.form()
    uploaded = []
    c = get_audit_conn()
    c.execute("""CREATE TABLE IF NOT EXISTS evidence_file (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT, uploaded_by TEXT, uploaded_at TEXT,
        pbc_ref TEXT, size INTEGER, data BLOB
    )""")
    for field_name, upload in form.multi_items():
        if hasattr(upload, "read"):
            data = await upload.read()
            if len(data) > MAX_MB * 1024 * 1024:
                continue
            c.execute(
                "INSERT INTO evidence_file (filename, uploaded_by, uploaded_at, size, data) VALUES (?,?,?,?,?)",
                (upload.filename, caller.get("username"), datetime.utcnow().isoformat(), len(data), data),
            )
            c.commit()
            fid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
            uploaded.append({"id": fid, "filename": upload.filename, "size": len(data)})
    c.close()
    _alog(request, caller, "evidence_upload", "evidence_file", payload={"count": len(uploaded)})
    return {"uploaded": uploaded}


# ── Audit Log ──────────────────────────────────────
@app.get("/api/audit/log")
async def audit_log_list(
    from_: str = Query(None, alias="from"), to: str = Query(None),
    user: str = Query(None), db: str = Query(None),
    limit: int = Query(50), offset: int = Query(0),
    verify: bool = Query(False),
    request: Request = None,
):
    caller = _caller(request)
    _audit_roles_required(caller, "admin", "audit_admin")
    c = get_audit_conn()

    if verify:
        result = verify_audit_log(c)
        c.close()
        return result

    where = ["1=1"]
    vals = []
    if from_: where.append("ts >= ?"); vals.append(from_)
    if to:    where.append("ts <= ?"); vals.append(to)
    if user:  where.append("username LIKE ?"); vals.append(f"%{user}%")
    if db:    where.append("db = ?"); vals.append(db)

    rows = c.execute(
        f"SELECT id, ts, username, action, target_type, target_id, db, ip FROM audit_log WHERE {' AND '.join(where)} ORDER BY id DESC LIMIT ? OFFSET ?",
        vals + [limit, offset],
    ).fetchall()
    c.close()
    cols = ["id","ts","username","action","target_type","target_id","db","ip"]
    return {"data": [dict(zip(cols, r)) for r in rows]}


# ── Package Export ─────────────────────────────────
@app.post("/api/audit/export/package")
async def audit_export_package(req: AuditPackageReq, request: Request):
    caller = _caller(request)
    _audit_roles_required(caller, *AUDIT_ROLES)
    c = get_audit_conn()

    buf = io.BytesIO()
    manifest = []

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        def add(name: str, content: str | bytes):
            data = content.encode() if isinstance(content, str) else content
            zf.writestr(name, data)
            manifest.append({
                "file": name,
                "sha256": hashlib.sha256(data).hexdigest(),
                "size": len(data),
            })

        # Engagement info
        eng = c.execute("SELECT * FROM audit_engagement WHERE id=?", (req.engagement_id,)).fetchone()
        if eng:
            cols = [d[1] for d in c.execute("PRAGMA table_info(audit_engagement)").fetchall()]
            add("engagement.json", json.dumps(dict(zip(cols, eng)), default=str, indent=2))

        # TB snapshot
        if req.include_tb:
            snaps = c.execute("SELECT * FROM tb_snapshot ORDER BY id DESC LIMIT 1").fetchall()
            for s in snaps:
                add(f"tb_snapshot_{s[2]}_{s[1]}.json", s[5])  # payload_json

        # Samples
        if req.include_samples:
            samples = c.execute("SELECT * FROM audit_sample WHERE engagement_id=?", (req.engagement_id,)).fetchall()
            cols = [d[1] for d in c.execute("PRAGMA table_info(audit_sample)").fetchall()]
            add("samples.json", json.dumps([dict(zip(cols, s)) for s in samples], default=str, indent=2))

        # Findings
        if req.include_findings:
            findings = c.execute("SELECT * FROM audit_finding WHERE engagement_id=?", (req.engagement_id,)).fetchall()
            cols = [d[1] for d in c.execute("PRAGMA table_info(audit_finding)").fetchall()]
            add("findings.json", json.dumps([dict(zip(cols, f)) for f in findings], default=str, indent=2))

        # PBC
        if req.include_pbc:
            pbc = c.execute("SELECT * FROM pbc_item WHERE engagement_id=?", (req.engagement_id,)).fetchall()
            cols = [d[1] for d in c.execute("PRAGMA table_info(pbc_item)").fetchall()]
            add("pbc_list.json", json.dumps([dict(zip(cols, p)) for p in pbc], default=str, indent=2))

        # Audit log extract
        log_rows = c.execute("SELECT id, ts, username, action, target_type, target_id, db FROM audit_log ORDER BY id DESC LIMIT 1000").fetchall()
        log_cols = ["id","ts","username","action","target_type","target_id","db"]
        add("audit_log.json", json.dumps([dict(zip(log_cols, r)) for r in log_rows], default=str, indent=2))

        # Manifest
        add("MANIFEST.json", json.dumps(manifest, indent=2))

    c.close()
    buf.seek(0)
    _alog(request, caller, "export_package", "engagement", req.engagement_id)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=audit-package-{req.engagement_id}.zip"},
    )
